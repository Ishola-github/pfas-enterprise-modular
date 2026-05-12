"""
SOP per-matrix pipeline orchestrator (PFAS Enterprise 5 SOP §6).

One canonical pipeline per matrix - never merged into one generalized pool.

Matrix             | Canonical dataset(s)            | pipeline_id
------------------ | ------------------------------- | -----------------
drinking water     | UCMR5                           | drinking_water
serum              | NHANES + SRM 1957               | serum
biosolids/sludge   | Method 1633 + EPA biosolids     | biosolids_sludge
AFFF               | RM 8690                         | afff
methanol standards | RM 8446                         | methanol_standards
air emissions      | OTM-50                          | air_emissions

The mapping above mirrors data/config/matrix_pipeline_sop.csv (single source of truth).

This orchestrator builds one matrix lane per invocation:

    python scripts/run_matrix_pipeline.py --lane drinking_water
    python scripts/run_matrix_pipeline.py --lane serum
    python scripts/run_matrix_pipeline.py --lane biosolids_sludge
    python scripts/run_matrix_pipeline.py --lane afff
    python scripts/run_matrix_pipeline.py --lane methanol_standards
    python scripts/run_matrix_pipeline.py --lane air_emissions

    python scripts/run_matrix_pipeline.py --lane all   # builds each lane separately

Each lane writes:

    data/training/<pipeline_id>/training.csv
    data/training/<pipeline_id>/manifest.json

Cross-matrix concatenation is explicitly NOT supported. The shared schema is for
*column compatibility*, not row pooling. The Shiny step 6) "Build multi-source
training table" (prepare_multisource_training.R) refuses to merge multiple
pipeline_lane sources into one CSV.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import re
import sys
import zipfile
from pathlib import Path
from typing import Any, Iterable

CANONICAL_OUTPUT_COLUMNS = [
    "source",
    "source_dataset",
    "source_file",
    "sample_id",
    "sample_name",
    "matrix",
    "sample_date",
    "analyte",
    "analyte_short_name",
    "cas_rn",
    "result_value_raw",
    "result_value_numeric",
    "qualifier",
    "result_unit",
    "mdl",
    "rl",
    "uncertainty",
    "coverage_factor",
    "method_id",
    "facility_id",
    "sample_point_id",
    "state",
    "county",
    "latitude",
    "longitude",
    "value_type",
    "matrix_governance_note",
    "pipeline_lane",
]


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _now_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _read_sop_lanes(project_root: Path) -> dict[str, dict[str, str]]:
    sop_path = project_root / "data" / "config" / "matrix_pipeline_sop.csv"
    if not sop_path.is_file():
        raise SystemExit(f"ERROR: missing SOP config: {sop_path}")
    out: dict[str, dict[str, str]] = {}
    with sop_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            pid = (row.get("pipeline_id") or "").strip()
            if pid:
                out[pid] = {k: (v or "").strip() for k, v in row.items()}
    return out


def _empty_row() -> dict[str, Any]:
    return {c: "" for c in CANONICAL_OUTPUT_COLUMNS}


def _write_training_csv(rows: list[dict[str, Any]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CANONICAL_OUTPUT_COLUMNS)
        writer.writeheader()
        for row in rows:
            full = _empty_row()
            for k, v in row.items():
                if k in full:
                    full[k] = "" if v is None else v
            writer.writerow(full)


def _write_manifest(
    project_root: Path,
    pipeline_id: str,
    *,
    sources: list[dict[str, Any]],
    out_csv: Path,
    rows_written: int,
    notes: list[str],
    sop_row: dict[str, str],
) -> Path:
    manifest_path = out_csv.parent / "manifest.json"
    payload = {
        "pipeline_id": pipeline_id,
        "matrix": sop_row.get("matrix", ""),
        "canonical_datasets": sop_row.get("canonical_datasets", ""),
        "generated_at_utc": _now_iso(),
        "training_csv": str(out_csv.relative_to(project_root)).replace("\\", "/"),
        "rows_written": rows_written,
        "sources": sources,
        "notes": notes,
        "sop_separation_note": (
            "SOP §6: this lane's rows are produced from its canonical dataset(s) only. "
            "Do not concatenate with rows from other pipeline_id lanes."
        ),
    }
    if rows_written > 0 and out_csv.is_file():
        payload["training_csv_sha256"] = _sha256_file(out_csv)
    manifest_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return manifest_path


def _coerce_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            f = float(value)
        except (TypeError, ValueError):
            return None
        if f != f:  # NaN
            return None
        return f
    s = str(value).strip()
    if not s:
        return None
    s_clean = s.replace(",", "")
    try:
        return float(s_clean)
    except ValueError:
        return None


# -------------------------------------------------------------------- #
# Lane builders                                                        #
# -------------------------------------------------------------------- #


def _build_drinking_water(project_root: Path, sop_row: dict[str, str]) -> dict[str, Any]:
    lane = "drinking_water"
    out_dir = project_root / "data" / "training" / lane
    out_csv = out_dir / "training.csv"
    sources: list[dict[str, Any]] = []
    notes: list[str] = []
    rows: list[dict[str, Any]] = []

    ucmr_csv = project_root / "data" / "training" / "ucmr_exceedance_labeled.csv"
    if not ucmr_csv.is_file():
        notes.append(
            "Missing data/training/ucmr_exceedance_labeled.csv — run scripts/run_ucmr_dataset_pipeline.R "
            "first (UCMR5 occurrence build)."
        )
        _write_training_csv(rows, out_csv)
        _write_manifest(
            project_root, lane,
            sources=sources, out_csv=out_csv, rows_written=0,
            notes=notes, sop_row=sop_row,
        )
        return {"pipeline_id": lane, "rows": 0, "csv": out_csv, "notes": notes}

    sources.append({"path": str(ucmr_csv.relative_to(project_root)).replace("\\", "/"),
                    "sha256": _sha256_file(ucmr_csv), "role": "UCMR5 finished-water occurrence (labeled)"})

    with ucmr_csv.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            analyte_name = (row.get("analyte_raw") or row.get("analyte") or row.get("Contaminant") or "").strip()
            analyte_short = (row.get("analyte_key") or row.get("analyte_short_name") or "").strip()
            cas_rn = (row.get("cas_key_effective") or row.get("casrn_ucmr")
                      or row.get("cas_rn") or row.get("CAS") or "").strip()
            sample_id = (row.get("sample_id") or row.get("SampleID") or "").strip()
            sample_date = (row.get("collection_date") or row.get("sample_date")
                           or row.get("CollectionDate") or "").strip()
            unit_raw = (row.get("unit_raw") or row.get("result_unit") or "").strip()
            facility_id = (row.get("pwsid") or row.get("facility_id") or row.get("PWSID") or "").strip()
            non_detect = (row.get("nondetect_flag") or "").strip()
            rv_raw = (row.get("result_raw") or row.get("result_value_raw")
                      or row.get("AnalyticalResultValue") or "").strip()
            ngl_raw = row.get("conc_ng_l")
            rv_num = _coerce_float(ngl_raw)
            if rv_num is None:
                rv_num = _coerce_float(row.get("result_num") or row.get("result_value_numeric")
                                       or row.get("AnalyticalResultValue"))
            qualifier = "ND" if non_detect == "1" else (row.get("qualifier") or "").strip()
            rows.append({
                "source": "EPA_UCMR5",
                "source_dataset": "UCMR5 (occurrence; ng/L harmonized)",
                "source_file": ucmr_csv.name,
                "sample_id": sample_id,
                "sample_name": (row.get("sample_name") or "").strip(),
                "matrix": "drinking water",
                "sample_date": sample_date,
                "analyte": analyte_name,
                "analyte_short_name": analyte_short,
                "cas_rn": cas_rn,
                "result_value_raw": rv_raw,
                "result_value_numeric": rv_num if rv_num is not None else "",
                "qualifier": qualifier,
                "result_unit": "ng/L",
                "mdl": (row.get("mdl") or "").strip(),
                "rl": (row.get("rl") or row.get("limit_ng_l") or "").strip(),
                "method_id": (row.get("method_id") or row.get("AnalyticalMethodID")
                              or "EPA_UCMR5_method").strip(),
                "facility_id": facility_id,
                "sample_point_id": (row.get("sample_point_id") or row.get("SamplePointID") or "").strip(),
                "state": (row.get("state") or row.get("State") or "").strip(),
                "value_type": "field_measurement",
                "matrix_governance_note": (
                    "EPA UCMR5 finished-water occurrence; matrix lane = drinking_water. "
                    f"unit_raw={unit_raw or 'NA'}; harmonized to ng/L via conc_ng_l. "
                    "Do not merge with serum, biosolids, AFFF, methanol, or air-emissions lanes."
                ),
                "pipeline_lane": lane,
            })

    _write_training_csv(rows, out_csv)
    _write_manifest(
        project_root, lane,
        sources=sources, out_csv=out_csv, rows_written=len(rows),
        notes=notes, sop_row=sop_row,
    )
    return {"pipeline_id": lane, "rows": len(rows), "csv": out_csv, "notes": notes}


def _build_serum(project_root: Path, sop_row: dict[str, str]) -> dict[str, Any]:
    lane = "serum"
    out_dir = project_root / "data" / "training" / lane
    out_csv = out_dir / "training.csv"
    sources: list[dict[str, Any]] = []
    notes: list[str] = []
    rows: list[dict[str, Any]] = []

    nh_proc = project_root / "data" / "processed" / "nhanes_pfas_with_demo.parquet"
    if nh_proc.is_file():
        sources.append({"path": str(nh_proc.relative_to(project_root)).replace("\\", "/"),
                        "sha256": _sha256_file(nh_proc),
                        "role": "NHANES PFAS + demographics (processed parquet)"})
        try:
            import pandas as pd
            df = pd.read_parquet(nh_proc)
            id_col = "SEQN" if "SEQN" in df.columns else df.columns[0]
            # NHANES P_PFAS coding: LBXPF* (e.g. LBXPFHS, LBXPFNA), LBXNFOA / LBXBFOA
            # (linear / branched PFOA), LBXNFOS / LBXMFOS (linear / monomethyl PFOS),
            # LBXMFOSAA (N-MeFOSAA), LBXMPAH (PFOSA derivative). Excludes LBD* (comments)
            # and LBXxxLC (limit-of-detection) flags.
            nhanes_pfas_codes = (
                "PFHS", "PFNA", "PFDE", "PFUA", "PFDOA",
                "NFOA", "BFOA", "NFOS", "MFOS", "MFOSAA",
                "MPAH",
            )
            pfas_cols = sorted({
                c for c in df.columns
                if c.startswith("LBX")
                and not c.endswith("LC")
                and any(code in c.upper() for code in nhanes_pfas_codes)
            })
            for _, r in df.iterrows():
                seqn = r.get(id_col)
                if seqn is None or (isinstance(seqn, float) and seqn != seqn):
                    sample_id_clean = ""
                else:
                    try:
                        sample_id_clean = str(int(float(seqn)))
                    except (TypeError, ValueError):
                        sample_id_clean = str(seqn)
                for col in pfas_cols:
                    val = r.get(col)
                    is_nan = isinstance(val, float) and val != val
                    rv_num = None if is_nan else _coerce_float(val)
                    raw_str = "" if (val is None or is_nan) else str(val)
                    rows.append({
                        "source": "CDC_NHANES",
                        "source_dataset": "NHANES PFAS (2017-2020 P_PFAS)",
                        "source_file": nh_proc.name,
                        "sample_id": sample_id_clean,
                        "sample_name": col,
                        "matrix": "serum",
                        "sample_date": "",
                        "analyte": col,
                        "analyte_short_name": col.replace("LBX", "").replace("PF", "PF"),
                        "cas_rn": "",
                        "result_value_raw": raw_str,
                        "result_value_numeric": rv_num if rv_num is not None else "",
                        "result_unit": "ng/mL",
                        "method_id": "CDC_NHANES_PFAS",
                        "value_type": "field_measurement",
                        "matrix_governance_note": (
                            "CDC NHANES human serum PFAS biomarker; matrix lane = serum. "
                            "Not drinking-water occurrence; not for MCL exceedance logic."
                        ),
                        "pipeline_lane": lane,
                    })
        except Exception as exc:  # noqa: BLE001
            notes.append(f"NHANES parquet read failed: {exc!r}")
    else:
        notes.append(
            "Missing data/processed/nhanes_pfas_with_demo.parquet — run "
            "scripts/train_nhanes_serum_pfas.py to produce it (NHANES XPT files are already present)."
        )

    srm_csv = project_root / "data" / "reference" / "nist" / "srm1957" / "serum_pfas.csv"
    if srm_csv.is_file():
        sources.append({"path": str(srm_csv.relative_to(project_root)).replace("\\", "/"),
                        "sha256": _sha256_file(srm_csv),
                        "role": "NIST SRM 1957 reconstituted human serum (Table A2 non-certified)"})
        with srm_csv.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rv_raw = (row.get("value") or "").strip()
                rows.append({
                    "source": "NIST_SRM1957",
                    "source_dataset": "NIST SRM 1957",
                    "source_file": srm_csv.name,
                    "sample_id": "NIST_SRM1957",
                    "sample_name": "SRM 1957",
                    "matrix": (row.get("matrix") or "serum").strip(),
                    "analyte": (row.get("analyte") or "").strip(),
                    "analyte_short_name": (row.get("short_name") or "").strip(),
                    "result_value_raw": rv_raw,
                    "result_value_numeric": _coerce_float(rv_raw) or "",
                    "uncertainty": (row.get("uncertainty") or "").strip(),
                    "coverage_factor": (row.get("coverage_factor") or "").strip(),
                    "result_unit": (row.get("unit") or "ug/kg").strip(),
                    "method_id": (row.get("method") or "LC-MS/MS").strip(),
                    "value_type": (row.get("value_type") or "non-certified").strip(),
                    "matrix_governance_note": (
                        "NIST SRM 1957 (serum) reference. Software benchmarking only; non-certified Table A2 values. "
                        "Do not merge with drinking-water, AFFF, methanol, or air-emissions lanes."
                    ),
                    "pipeline_lane": lane,
                })
    else:
        notes.append("Missing data/reference/nist/srm1957/serum_pfas.csv (NIST SRM 1957 reference).")

    _write_training_csv(rows, out_csv)
    _write_manifest(
        project_root, lane,
        sources=sources, out_csv=out_csv, rows_written=len(rows),
        notes=notes, sop_row=sop_row,
    )
    return {"pipeline_id": lane, "rows": len(rows), "csv": out_csv, "notes": notes}


def _build_biosolids_sludge(project_root: Path, sop_row: dict[str, str]) -> dict[str, Any]:
    """Biosolids/sludge lane.

    Canonical facility universe: NPDES_BIOSOLIDS_PERMITS.csv (one row per
    biosolids-permitted facility). Enriched with per-facility counts from
    NPDES_BIOSOLIDS_INSPECTIONS, NPDES_BIOSOLIDS_SEV_VIOLATIONS,
    NPDES_BIOSOLIDS_FORMAL_ACTIONS, and NPDES_BIOSOLIDS_INFML_ENF_ACTIONS.

    The training rows from this lane are program metadata (facility universe +
    compliance context), NOT nationwide PFAS sludge concentrations. EPA does
    not publish per-facility PFAS-in-biosolids measurements as a bulk export;
    PFAS analytical context for biosolids/sludge comes from EPA Method 1633(A).
    Downstream code must treat these rows as the *biosolids matrix label space*,
    not as analytical results.
    """

    lane = "biosolids_sludge"
    out_dir = project_root / "data" / "training" / lane
    out_csv = out_dir / "training.csv"
    sources: list[dict[str, Any]] = []
    notes: list[str] = []
    rows: list[dict[str, Any]] = []

    method_meta = project_root / "data" / "reference" / "epa_1633a_method_metadata.csv"
    method_metadata: list[dict[str, str]] = []
    if method_meta.is_file():
        sources.append({"path": str(method_meta.relative_to(project_root)).replace("\\", "/"),
                        "sha256": _sha256_file(method_meta),
                        "role": "EPA Method 1633(A) metadata (analyte / LOD / LOQ governance)"})
        with method_meta.open("r", encoding="utf-8", newline="") as f:
            method_metadata = [dict(r) for r in csv.DictReader(f)]
    else:
        notes.append("Missing data/reference/epa_1633a_method_metadata.csv (EPA Method 1633A metadata).")

    biosolids_zip = project_root / "data" / "raw" / "epa_icis_npdes" / "npdes_biosolids_downloads.zip"
    permits_seen = 0
    biosolids_flag_y = 0

    if not biosolids_zip.is_file():
        notes.append(
            "Missing data/raw/epa_icis_npdes/npdes_biosolids_downloads.zip — run "
            ".\\download_epa_icis_npdes_ml.ps1 -SkipDmr -SkipOutfalls to fetch the biosolids ZIP."
        )
    else:
        sources.append({"path": str(biosolids_zip.relative_to(project_root)).replace("\\", "/"),
                        "sha256": _sha256_file(biosolids_zip),
                        "role": "EPA ECHO ICIS-NPDES biosolids facility / permit / compliance bulk export (ZIP)"})
        try:
            with zipfile.ZipFile(biosolids_zip) as z:
                members = {Path(n).name.upper(): n for n in z.namelist() if n.upper().endswith(".CSV")}

                def _read_member(upper_name: str) -> tuple[str, list[dict[str, str]]] | None:
                    member = members.get(upper_name)
                    if not member:
                        return None
                    with z.open(member) as fh:
                        text = fh.read().decode("latin-1", errors="replace")
                    return Path(member).name, list(csv.DictReader(text.splitlines()))

                def _count_by(rows_in: list[dict[str, str]], keys: tuple[str, ...]) -> dict[str, int]:
                    out: dict[str, int] = {}
                    for r in rows_in:
                        key_val = ""
                        for k in keys:
                            v = (r.get(k) or "").strip()
                            if v:
                                key_val = v
                                break
                        if key_val:
                            out[key_val] = out.get(key_val, 0) + 1
                    return out

                permits = _read_member("NPDES_BIOSOLIDS_PERMITS.CSV")
                inspections = _read_member("NPDES_BIOSOLIDS_INSPECTIONS.CSV")
                sev_viol = _read_member("NPDES_BIOSOLIDS_SEV_VIOLATIONS.CSV")
                formal = _read_member("NPDES_BIOSOLIDS_FORMAL_ACTIONS.CSV")
                informal = _read_member("NPDES_BIOSOLIDS_INFML_ENF_ACTIONS.CSV")

                inspections_by_id = _count_by(inspections[1] if inspections else [], ("NPDES_ID",))
                sev_by_id = _count_by(sev_viol[1] if sev_viol else [], ("SOURCE_ID", "NPDES_ID"))
                formal_by_id = _count_by(formal[1] if formal else [], ("NPDES_ID",))
                informal_by_id = _count_by(informal[1] if informal else [], ("NPDES_ID",))

                if not permits:
                    notes.append(f"NPDES_BIOSOLIDS_PERMITS.csv not present in {biosolids_zip.name}; "
                                 f"members={list(members.keys())[:8]}")
                else:
                    permits_file, permits_rows = permits

                    for raw in permits_rows:
                        permits_seen += 1
                        npdes_id = (raw.get("NPDES_ID") or "").strip()
                        if not npdes_id:
                            continue
                        bflag = (raw.get("BIOSOLIDS_FLAG") or "").strip().upper()
                        if bflag != "Y":
                            continue
                        biosolids_flag_y += 1

                        cwp_name = (raw.get("CWP_NAME") or "").strip()
                        state_val = (raw.get("CWP_STATE") or "").strip()
                        county_val = (raw.get("CWP_COUNTY") or "").strip()
                        permit_status = (raw.get("CWP_PERMIT_STATUS_DESC") or "").strip()
                        facility_type = (raw.get("CWP_FACILITY_TYPE_INDICATOR") or "").strip()

                        ins_n = inspections_by_id.get(npdes_id, 0)
                        sev_n = sev_by_id.get(npdes_id, 0)
                        fa_n = formal_by_id.get(npdes_id, 0)
                        ifa_n = informal_by_id.get(npdes_id, 0)

                        gov_note = (
                            "ICIS-NPDES biosolids program metadata (facility universe). "
                            "PFAS analytical context = EPA Method 1633(A); this row is NOT a PFAS measurement. "
                            f"permit_status={permit_status or 'NA'}; facility_type={facility_type or 'NA'}; "
                            f"inspections={ins_n}; severe_violations={sev_n}; "
                            f"formal_actions={fa_n}; informal_actions={ifa_n}."
                        )

                        rows.append({
                            "source": "EPA_ICIS_NPDES_BIOSOLIDS",
                            "source_dataset": "EPA ICIS-NPDES biosolids permits (BIOSOLIDS_FLAG=Y)",
                            "source_file": permits_file,
                            "sample_id": npdes_id,
                            "sample_name": cwp_name or npdes_id,
                            "matrix": "biosolids/sludge",
                            "facility_id": npdes_id,
                            "state": state_val,
                            "county": county_val,
                            "method_id": "EPA_1633A_metadata",
                            "value_type": "program_metadata",
                            "matrix_governance_note": gov_note,
                            "pipeline_lane": lane,
                        })

                    for label, table in [
                        ("inspections", inspections),
                        ("sev_violations", sev_viol),
                        ("formal_actions", formal),
                        ("informal_actions", informal),
                    ]:
                        if table is None:
                            notes.append(f"NPDES_BIOSOLIDS_{label.upper()}.csv not present in ZIP.")
                        else:
                            fname, recs = table
                            notes.append(f"{fname}: {len(recs)} records (used for per-facility enrichment).")

        except zipfile.BadZipFile:
            notes.append(f"npdes_biosolids_downloads.zip is not a valid zip file: {biosolids_zip}")

    if method_metadata and rows:
        first_meta = method_metadata[0]
        rows.append({
            "source": "EPA_METHOD_1633A",
            "source_dataset": "EPA Method 1633(A) metadata",
            "source_file": method_meta.name,
            "sample_id": "method_1633a_metadata",
            "matrix": "biosolids/sludge",
            "analyte": first_meta.get("analyte", "") or first_meta.get("Analyte", ""),
            "result_unit": first_meta.get("unit", "") or first_meta.get("Unit", "ng/g"),
            "method_id": "EPA_1633A",
            "value_type": "method_metadata",
            "matrix_governance_note": (
                "EPA Method 1633A metadata for biosolids/sludge analytical chemistry; not a measurement row."
            ),
            "pipeline_lane": lane,
        })

    if permits_seen:
        notes.append(
            f"NPDES_BIOSOLIDS_PERMITS.csv: {permits_seen} permits in file, "
            f"{biosolids_flag_y} retained (BIOSOLIDS_FLAG='Y')."
        )
    notes.append(
        "SCOPE: this lane emits program/facility metadata + EPA Method 1633A analytical metadata only. "
        "It does NOT contain nationwide PFAS-in-biosolids concentrations (no such bulk dataset is published by EPA). "
        "Use upstream Method 1633(A) PFAS sludge measurements (state programs, EPA studies, lab reports) for "
        "analytical training data; treat these rows as the matrix label space + compliance enrichment."
    )

    _write_training_csv(rows, out_csv)
    _write_manifest(
        project_root, lane,
        sources=sources, out_csv=out_csv, rows_written=len(rows),
        notes=notes, sop_row=sop_row,
    )
    return {"pipeline_id": lane, "rows": len(rows), "csv": out_csv, "notes": notes}


def _build_nist_reference_lane(
    project_root: Path,
    sop_row: dict[str, str],
    *,
    pipeline_id: str,
    csv_rel: str,
    canonical_unit_fallback: str,
    governance_note: str,
) -> dict[str, Any]:
    out_dir = project_root / "data" / "training" / pipeline_id
    out_csv = out_dir / "training.csv"
    sources: list[dict[str, Any]] = []
    notes: list[str] = []
    rows: list[dict[str, Any]] = []

    src_csv = project_root / csv_rel
    if not src_csv.is_file():
        notes.append(f"Missing canonical reference CSV: {csv_rel}")
    else:
        sources.append({"path": csv_rel.replace("\\", "/"),
                        "sha256": _sha256_file(src_csv),
                        "role": f"NIST reference table for {pipeline_id} lane"})
        with src_csv.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rv_raw = (row.get("value") or "").strip()
                rows.append({
                    "source": f"NIST_{(row.get('reference_material') or '').replace(' ', '')}",
                    "source_dataset": (row.get("reference_material") or "").strip(),
                    "source_file": src_csv.name,
                    "sample_id": (row.get("reference_material") or "").replace(" ", "_"),
                    "sample_name": (row.get("reference_material") or "").strip(),
                    "matrix": (row.get("matrix") or "").strip(),
                    "analyte": (row.get("analyte") or "").strip(),
                    "analyte_short_name": (row.get("short_name") or "").strip(),
                    "result_value_raw": rv_raw,
                    "result_value_numeric": _coerce_float(rv_raw) or "",
                    "uncertainty": (row.get("uncertainty") or "").strip(),
                    "coverage_factor": (row.get("coverage_factor") or "").strip(),
                    "result_unit": (row.get("unit") or canonical_unit_fallback).strip(),
                    "method_id": (row.get("method") or "LC-MS/MS").strip(),
                    "value_type": (row.get("value_type") or "non-certified").strip(),
                    "matrix_governance_note": governance_note,
                    "pipeline_lane": pipeline_id,
                })

    _write_training_csv(rows, out_csv)
    _write_manifest(
        project_root, pipeline_id,
        sources=sources, out_csv=out_csv, rows_written=len(rows),
        notes=notes, sop_row=sop_row,
    )
    return {"pipeline_id": pipeline_id, "rows": len(rows), "csv": out_csv, "notes": notes}


def _build_afff(project_root: Path, sop_row: dict[str, str]) -> dict[str, Any]:
    return _build_nist_reference_lane(
        project_root, sop_row,
        pipeline_id="afff",
        csv_rel="data/reference/nist/rm8690/afff_pfas.csv",
        canonical_unit_fallback="ug/g",
        governance_note=(
            "NIST RM 8690 AFFF foam reference; foam/forensic lane. "
            "Do not merge with drinking-water (UCMR), serum (NHANES/SRM1957), biosolids, methanol, or air-emissions lanes."
        ),
    )


def _build_methanol_standards(project_root: Path, sop_row: dict[str, str]) -> dict[str, Any]:
    return _build_nist_reference_lane(
        project_root, sop_row,
        pipeline_id="methanol_standards",
        csv_rel="data/reference/nist/rm8446/methanol_pfas.csv",
        canonical_unit_fallback="mg/kg",
        governance_note=(
            "NIST RM 8446 methanol calibration-style reference. "
            "Not an environmental matrix; do not treat as a field sample. "
            "Do not merge with drinking-water, serum, biosolids, AFFF, or air-emissions lanes."
        ),
    )


def _build_air_emissions(project_root: Path, sop_row: dict[str, str]) -> dict[str, Any]:
    lane = "air_emissions"
    out_dir = project_root / "data" / "training" / lane
    out_csv = out_dir / "training.csv"
    sources: list[dict[str, Any]] = []
    notes: list[str] = []
    rows: list[dict[str, Any]] = []

    otm_dir = project_root / "data" / "external" / "epa_otm50"
    workbooks = sorted(otm_dir.glob("*.xlsx")) if otm_dir.is_dir() else []
    if not workbooks:
        notes.append(
            "Missing OTM-50 XLSX workbooks under data/external/epa_otm50/ — run .\\download_epa_otm50.ps1."
        )
        _write_training_csv(rows, out_csv)
        _write_manifest(
            project_root, lane,
            sources=sources, out_csv=out_csv, rows_written=0,
            notes=notes, sop_row=sop_row,
        )
        return {"pipeline_id": lane, "rows": 0, "csv": out_csv, "notes": notes}

    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        notes.append(f"openpyxl is required for OTM-50: {exc!r}")
        _write_training_csv(rows, out_csv)
        _write_manifest(
            project_root, lane,
            sources=sources, out_csv=out_csv, rows_written=0,
            notes=notes, sop_row=sop_row,
        )
        return {"pipeline_id": lane, "rows": 0, "csv": out_csv, "notes": notes}

    for wb_path in workbooks:
        sources.append({
            "path": str(wb_path.relative_to(project_root)).replace("\\", "/"),
            "sha256": _sha256_file(wb_path),
            "role": "EPA OTM-50 fluoropolymer manufacturer air-emissions workbook",
        })
        wb = openpyxl.load_workbook(wb_path, data_only=True)
        for sheet_name in wb.sheetnames:
            low = sheet_name.lower()
            if "readme" in low or "log" in low or "login" in low:
                continue
            ws = wb[sheet_name]
            data = list(ws.iter_rows(values_only=True))
            if not data:
                continue
            cas_row_idx = None
            for i, row in enumerate(data):
                if row is None:
                    continue
                if any((isinstance(c, str) and c.strip().upper() in {"CAS #", "CAS#", "CAS RN", "CAS"}) for c in row):
                    cas_row_idx = i
                    break
            if cas_row_idx is None:
                notes.append(f"{wb_path.name}::{sheet_name}: no CAS header row found, skipped.")
                continue
            sample_header_idx = max(0, cas_row_idx - 1)
            unit_row_idx = cas_row_idx + 1 if cas_row_idx + 1 < len(data) else cas_row_idx
            sample_header = data[sample_header_idx] or ()
            cas_row = data[cas_row_idx] or ()
            unit_row = data[unit_row_idx] or ()
            n_cols = len(cas_row)
            mdl_col = None
            rl_col = None
            for j in range(n_cols):
                cell = sample_header[j] if j < len(sample_header) else None
                if isinstance(cell, str):
                    s = cell.strip().upper()
                    if s == "MDL":
                        mdl_col = j
                    elif s == "RL":
                        rl_col = j
                cell2 = cas_row[j] if j < len(cas_row) else None
                if isinstance(cell2, str):
                    s2 = cell2.strip().upper()
                    if s2 == "MDL" and mdl_col is None:
                        mdl_col = j
                    elif s2 == "RL" and rl_col is None:
                        rl_col = j
            sample_cols: list[int] = []
            for j in range(2, n_cols):
                if j in {mdl_col, rl_col}:
                    continue
                hdr = sample_header[j] if j < len(sample_header) else None
                cas_hdr = cas_row[j] if j < len(cas_row) else None
                if hdr is None and cas_hdr is None:
                    continue
                if isinstance(hdr, str) and hdr.strip().lower() in {"conversion"}:
                    continue
                sample_cols.append(j)
            for r_i in range(cas_row_idx + 2, len(data)):
                row = data[r_i]
                if row is None:
                    continue
                analyte = row[0] if len(row) > 0 else None
                cas_rn = row[1] if len(row) > 1 else None
                if not isinstance(analyte, str) or not analyte.strip():
                    continue
                if analyte.strip().lower().startswith("note") or analyte.strip().lower().startswith("comment"):
                    continue
                analyte_name = analyte.strip()
                cas_clean = str(cas_rn).strip() if cas_rn is not None else ""
                mdl_val = row[mdl_col] if (mdl_col is not None and mdl_col < len(row)) else None
                rl_val = row[rl_col] if (rl_col is not None and rl_col < len(row)) else None
                for j in sample_cols:
                    if j >= len(row):
                        continue
                    val = row[j]
                    if val is None:
                        continue
                    hdr = sample_header[j] if j < len(sample_header) else ""
                    canister_id = cas_row[j] if j < len(cas_row) else ""
                    unit = unit_row[j] if j < len(unit_row) else None
                    unit_clean = "ug/m3"
                    if isinstance(unit, str):
                        u = unit.strip().replace("\ufffd", "u").replace("µ", "u").lower()
                        if u:
                            unit_clean = u.replace("um3", "ug/m3").replace("u g/m3", "ug/m3")
                    raw_str = str(val).strip()
                    qualifier = ""
                    rv_num = _coerce_float(val)
                    if rv_num is None:
                        if raw_str.upper() == "BDL":
                            qualifier = "BDL"
                    rows.append({
                        "source": "EPA_OTM50",
                        "source_dataset": "EPA OTM-50 (Chemours Fayetteville Works)",
                        "source_file": wb_path.name,
                        "sample_id": f"{wb_path.stem}::{canister_id}" if canister_id else f"{wb_path.stem}::{hdr}",
                        "sample_name": str(hdr).strip() if hdr is not None else "",
                        "matrix": "air_emissions",
                        "analyte": analyte_name,
                        "cas_rn": cas_clean,
                        "result_value_raw": raw_str,
                        "result_value_numeric": rv_num if rv_num is not None else "",
                        "qualifier": qualifier,
                        "result_unit": unit_clean,
                        "mdl": _coerce_float(mdl_val) if mdl_val is not None else "",
                        "rl": _coerce_float(rl_val) if rl_val is not None else "",
                        "method_id": "EPA_OTM50",
                        "value_type": "field_measurement",
                        "matrix_governance_note": (
                            "EPA OTM-50 stack-gas / process-control PFAS air emissions at a single fluoropolymer manufacturer. "
                            "Air-emissions matrix only; do not merge with UCMR (drinking water), NHANES/SRM1957 (serum), "
                            "ICIS-NPDES (biosolids/wastewater), RM 8690 (AFFF), or RM 8446 (methanol) lanes."
                        ),
                        "pipeline_lane": lane,
                    })

    _write_training_csv(rows, out_csv)
    _write_manifest(
        project_root, lane,
        sources=sources, out_csv=out_csv, rows_written=len(rows),
        notes=notes, sop_row=sop_row,
    )
    return {"pipeline_id": lane, "rows": len(rows), "csv": out_csv, "notes": notes}


# -------------------------------------------------------------------- #
# Dispatcher                                                           #
# -------------------------------------------------------------------- #

BUILDERS = {
    "drinking_water": _build_drinking_water,
    "serum": _build_serum,
    "biosolids_sludge": _build_biosolids_sludge,
    "afff": _build_afff,
    "methanol_standards": _build_methanol_standards,
    "air_emissions": _build_air_emissions,
}


def run_lane(project_root: Path, lane: str, sop_lanes: dict[str, dict[str, str]]) -> dict[str, Any]:
    if lane not in BUILDERS:
        raise SystemExit(
            f"ERROR: unknown pipeline lane {lane!r}. "
            f"Valid lanes: {sorted(BUILDERS.keys())}"
        )
    if lane not in sop_lanes:
        raise SystemExit(
            f"ERROR: lane {lane!r} not in data/config/matrix_pipeline_sop.csv. "
            "Add it to the SOP first."
        )
    builder = BUILDERS[lane]
    return builder(project_root, sop_lanes[lane])


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument(
        "--lane",
        type=str,
        required=True,
        help=("Pipeline lane to build. One of: "
              "drinking_water, serum, biosolids_sludge, afff, methanol_standards, air_emissions, all."),
    )
    parser.add_argument(
        "--project-root",
        type=Path,
        default=Path(__file__).resolve().parent.parent,
        help="Repository root (default: parent of scripts/).",
    )
    args = parser.parse_args()

    project_root = args.project_root.resolve()
    sop_lanes = _read_sop_lanes(project_root)
    lane = args.lane.strip().lower()

    if lane == "all":
        results: list[dict[str, Any]] = []
        for pid in BUILDERS:
            print(f"\n=== Building lane: {pid} ===")
            result = run_lane(project_root, pid, sop_lanes)
            print(f"  rows={result['rows']}  csv={result['csv']}")
            if result["notes"]:
                for n in result["notes"]:
                    print(f"  note: {n}")
            results.append(result)
        # Cross-lane summary; no concatenation.
        summary_path = project_root / "data" / "training" / "matrix_pipeline_summary.json"
        summary = {
            "generated_at_utc": _now_iso(),
            "sop_separation_note": (
                "Each lane has its own training.csv and manifest.json. "
                "This summary lists per-lane outputs; rows from different lanes are NEVER merged here."
            ),
            "lanes": [
                {
                    "pipeline_id": r["pipeline_id"],
                    "rows": r["rows"],
                    "csv": str(r["csv"].relative_to(project_root)).replace("\\", "/"),
                    "notes": r["notes"],
                }
                for r in results
            ],
        }
        summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
        print(f"\nWrote cross-lane summary: {summary_path.relative_to(project_root)}")
        return 0

    result = run_lane(project_root, lane, sop_lanes)
    print(f"lane={result['pipeline_id']}  rows={result['rows']}  csv={result['csv']}")
    if result["notes"]:
        for n in result["notes"]:
            print(f"note: {n}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
